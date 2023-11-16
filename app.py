from flask import Flask,render_template,request,jsonify
from flask_sqlalchemy import SQLAlchemy
from urllib.parse import quote_plus
import pandas as pd
import openpyxl

app=Flask(__name__)

username='postgres'
password='123456'
host='@localhost'
port=5432


app.config['SQLALCHEMY_DATABASE_URI'] = f'postgresql://{username}:{password}{host}:{port}/neet'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

class Clg(db.Model):
    id = db.Column(db.Integer(), primary_key=True,autoincrement=True)
    college_code = db.Column(db.String(100),nullable=False)
    college_name = db.Column(db.String(100),nullable=False)
    city = db.Column(db.String(100),nullable=False)
    state = db.Column(db.String(100),nullable=False)
    course_name = db.Column(db.String(100),nullable=False)
    category = db.Column(db.String(100),nullable=False)
    no_of_seats = db.Column(db.String(100),nullable=False)

    def __repr__(self):
        return f'{self.college_code} ok'
    
class Detail(db.Model):
    id = db.Column(db.Integer, primary_key=True,autoincrement=True)
    college_name = db.Column(db.String(200), nullable=False)
    course_name = db.Column(db.String(100), nullable=False)
    category = db.Column(db.String(100),nullable=False)
    rank_high = db.Column(db.String(100), nullable=False)
    rank_low = db.Column(db.String(100), nullable=False)
    marks_high = db.Column(db.String(100), nullable=False)
    marks_low = db.Column(db.String(100), nullable=False)
    period = db.Column(db.String(100), nullable=False)
    

@app.route('/')
def home():
    a=Clg.query.all()
    return render_template('index.html',a=a)



@app.route('/new',methods=['POST'])
def college():
    if request.method=='POST':
        file=request.files['college']
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        path= file
        df = pd.read_excel(path)
        # print(df)
        num=len(df)+1
        a=[]
        for row in sheet.iter_rows(min_row=2, max_row=num, values_only=True):
            a.append(row)

        for id,college_code,college_name,city,state,course_name,category,no_of_seats in a:
            data = Clg(id=id,college_code=college_code,college_name=college_name,city=city,state=state,course_name=course_name,category=category,no_of_seats=no_of_seats)
            db.session.add(data)
        db.session.commit()
    return jsonify({'msg':'College data entered successfully'})

@app.route('/',methods=['POST'])
def choosefile():
    if request.method=='POST':
        file=request.files['file']
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        path= file
        df = pd.read_excel(path)
        # print(df)
        num=len(df)+1
        a=[]
        for row in sheet.iter_rows(min_row=2, max_row=num, values_only=True):
            a.append(row)

        for id,college_name,course_name,category,rank_high,rank_low,marks_high,marks_low,period in a:
            data = Detail(id=id,college_name=college_name,course_name=course_name,category=category,rank_high=rank_high,rank_low=rank_low,marks_high=marks_high,marks_low=marks_low,period=period)
            db.session.add(data)
        db.session.commit()
    return jsonify({'msg':'cutoff data entered successfully '})

if __name__ == '__main__':
    app.run(debug=True)